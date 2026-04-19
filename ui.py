import os
import collections
import ctypes
import platform
import sys
import threading
import time
import asyncio
import concurrent.futures
import csv
import tkinter as tk
from ctypes import wintypes
from pathlib import Path
from typing import Callable

try:
    import tkinterdnd2 as _tkdnd
    _TK_DND_AVAILABLE = True
except ImportError:
    _tkdnd = None
    _TK_DND_AVAILABLE = False

from ai_bridge import AIBridge
from database import Database
try:
    from live_listen import LiveInterviewListener
except Exception:
    LiveInterviewListener = None

WDA_MONITOR = 1
WDA_EXCLUDEFROMCAPTURE = 0x00000011


class OverlayApp:
    """Tkinter overlay UI for the assistant.

    This class focuses on UI concerns and delegates AI calls to `AIBridge`.
    """

    def __init__(self):
        if _TK_DND_AVAILABLE:
            try:
                self.root = _tkdnd.TkinterDnD.Tk()
                self._dnd_available = True
            except Exception:
                self.root = tk.Tk()
                self._dnd_available = False
        else:
            self.root = tk.Tk()
            self._dnd_available = False
        self.root.title("Stealth ChatGPT Overlay")
        self._capture_affinity = None
        self._capture_error_logged = False
        self.root.overrideredirect(True)
        self.root.attributes("-topmost", True)
        self.root.attributes("-alpha", 0.7)
        self.root.geometry("640x320+100+100")
        self.root.configure(bg="#111111")

        self.drag_offset = (0, 0)

        # async loop and AI bridge for fast streaming responses
        self.db = Database()
        self.loop = asyncio.new_event_loop()
        threading.Thread(target=self._start_async_loop, daemon=True).start()
        self.ai = AIBridge(self.db)
        self.selected_model = "gpt-4.1"
        # pre-warm the AI HTTP session to reduce first-request latency
        try:
            asyncio.run_coroutine_threadsafe(self.ai.warmup(), self.loop)
        except Exception:
            pass

        # interview listener (created on demand)
        self._listener = None

        self._build_ui()
        # apply taskbar-hide styles shortly after UI is built (Windows only)
        try:
            self.root.after(50, self._apply_taskbar_hide)
        except Exception:
            pass
        self.root.after(200, self._hide_from_capture)
        self.root.protocol("WM_DELETE_WINDOW", self.stop)

    def _start_async_loop(self):
        asyncio.set_event_loop(self.loop)
        self.loop.run_forever()

    def _build_ui(self):
        # small draggable header to save vertical space (title/status hidden)
        header_frame = tk.Frame(self.root, height=16, bg="#111111", cursor="arrow")
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        grip = tk.Label(header_frame, text=" ", bg="#111111")
        grip.pack(side="left", padx=2, pady=1)
        # header controls: camera (capture) and a close (X) button
        # pack close first so camera appears to its left
        self.header_close = tk.Button(
            header_frame,
            text="✕",
            command=self.stop,
            bg="#333333",
            fg="white",
            relief="flat",
            padx=6,
            pady=0,
            bd=0,
            cursor="arrow"
        )
        self.header_close.pack(side="right", padx=6, pady=2)

        self.header_capture = tk.Button(
            header_frame,
            text="📸",
            command=self.capture_screen_and_answer,
            bg="#e69900",
            fg="black",
            relief="flat",
            padx=6,
            pady=0,
            bd=0,
            cursor="arrow"
        )
        self.header_capture.pack(side="right", padx=6, pady=2)

        # status_label kept for programmatic updates but not displayed to save space
        self.status_label = tk.Label(
            self.root,
            text="",
            fg="#a6e22e",
            bg="#111111",
            font=("Segoe UI", 9)
        )

        log_frame = tk.Frame(self.root, bg="#111111")
        log_frame.pack(padx=6, fill="both", expand=True)

        self.log_scrollbar = tk.Scrollbar(log_frame, orient="vertical", cursor="arrow")
        self.log_scrollbar.pack(side="right", fill="y")

        self.log = tk.Text(
            log_frame,
            bg="#111111",
            fg="#ffffff",
            insertbackground="#ffffff",
            selectbackground="#44475a",
            selectforeground="#f8f8f2",
            exportselection=True,
            takefocus=True,
            cursor="arrow",
            height=13,
            relief="flat",
            wrap="word",
            yscrollcommand=self.log_scrollbar.set
        )
        self.log_scrollbar.config(command=self.log.yview)
        # slightly smaller default font for denser display
        try:
            self.log.configure(font=("Segoe UI", 9))
        except Exception:
            pass
        # reduce inter-line spacing to make the log denser; darken assistant yellow
        self.log.tag_configure("user", foreground="#b8860b", background="#1b1d2a", spacing1=1, spacing3=2)
        self.log.tag_configure("assistant", foreground="#50fa7b", background="#282a36", spacing1=1, spacing3=2)
        self.log.tag_configure("system", foreground="#50fa7b", background="#111111", spacing1=1, spacing3=2)
        self.log.tag_configure("thinking", foreground="#888888", background="#1a1a1a", spacing1=1, spacing3=2)
        self.log.pack(side="left", fill="both", expand=True)
        self.log.bind("<Control-c>", self.copy_selection)
        self.log.bind("<Control-Insert>", self.copy_selection)
        self.log.bind("<Control-a>", self.select_all)
        self.log.bind("<Button-3>", self._show_context_menu)
        # detect user scroll/interaction to stop auto-anchoring
        self.log.bind("<MouseWheel>", self._on_user_scroll)
        self.log.bind("<Button-4>", self._on_user_scroll)
        self.log.bind("<Button-5>", self._on_user_scroll)
        self.log.bind("<Button-1>", self._on_user_scroll)

        self.text_menu = tk.Menu(self.root, tearoff=0)
        self.text_menu.add_command(label="Copy", command=self.copy_selection)
        self.text_menu.add_command(label="Select All", command=self.select_all)

        prompt_frame = tk.Frame(self.root, bg="#111111")
        prompt_frame.pack(fill="x", padx=8, pady=(2, 6))

        self.prompt_entry = tk.Entry(
            prompt_frame,
            bg="#222222",
            fg="white",
            insertbackground="white",
            relief="flat",
            font=("Segoe UI", 9),
            cursor="arrow"
        )
        self.prompt_entry.pack(side="left", fill="x", expand=True, ipady=2)
        self.prompt_entry.bind("<Return>", self.send_prompt)

        self._attach_button = tk.Button(
            prompt_frame,
            text="📎",
            command=self._browse_file,
            bg="#333333",
            fg="white",
            relief="flat",
            padx=5,
            pady=0,
            bd=0,
            cursor="arrow",
        )
        self._attach_button.pack(side="left", padx=(4, 0))

        # language selection dropdown (default: Python)
        self.lang_var = tk.StringVar(value="Python")
        lang_options = ["Python", "TypeScript", "Java", "JavaScript", "C++", "C#", "Go", "Ruby"]
        try:
            lang_menu = tk.OptionMenu(prompt_frame, self.lang_var, *lang_options, command=self._on_lang_change)
            lang_menu.config(bg="#222222", fg="white", relief="flat")
            lang_menu.pack(side="left", padx=6)
        except Exception:
            # fallback: no dropdown on environments with issues
            pass

        model_frame = tk.Frame(prompt_frame, bg="#111111")
        model_frame.pack(side="left", padx=(0, 6))

        self.gpt4o_button = tk.Button(
            model_frame,
            text="GPT-4.1",
            command=lambda: self._set_selected_model("gpt-4.1"),
            width=8,
            relief="flat",
        )
        self.gpt4o_button.pack(side="left", padx=(0, 4))

        self.gpt5_button = tk.Button(
            model_frame,
            text="GPT-5",
            command=lambda: self._set_selected_model("gpt-5"),
            width=8,
            relief="flat",
        )
        self.gpt5_button.pack(side="left")

        self._refresh_model_buttons()
        

        self.send_button = tk.Button(
            prompt_frame,
            text="Send",
            command=self.send_prompt,
            width=10,
            bg="#44b94e",
            fg="white",
            relief="flat"
        )
        self.send_button.pack(side="left", padx=(6, 0))

        self.stop_button = tk.Button(
            prompt_frame,
            text="Stop",
            command=self._handle_stop_action,
            width=10,
            bg="#b94e44",
            fg="white",
            relief="flat",
            state="disabled"
        )
        self.stop_button.pack(side="left", padx=(6, 0))
        self.cancel_send_button = self.stop_button

        # Mic toggle for interview listening (non-disruptive)
        try:
            self.listener_button = tk.Button(
                prompt_frame,
                text="Mic",
                command=self._toggle_listener,
                width=8,
                bg="#333333",
                fg="white",
                relief="flat"
            )
            self.listener_button.pack(side="left", padx=6)
            # compact status label — shows rms/flushing without touching the log
            self._listener_status_lbl = tk.Label(
                prompt_frame,
                text="",
                bg="#111111",
                fg="#666666",
                font=("Consolas", 8),
                anchor="w",
            )
            self._listener_status_lbl.pack(side="left", padx=(0, 6), fill="x")
        except Exception:
            pass

        # file attachment row — always visible as drop zone
        self._file_row = tk.Frame(self.root, bg="#1a1a2e", pady=2)
        self._file_row.pack(fill="x", padx=8)
        self._file_indicator_label = tk.Label(
            self._file_row,
            text="  Drop file here  ·  Ctrl+V  ·  📎  ",
            bg="#1a1a2e",
            fg="#555577",
            font=("Segoe UI", 8),
            anchor="w",
            padx=4,
            cursor="arrow",
        )
        self._file_indicator_label.pack(side="left", fill="x", expand=True, padx=(4, 0), pady=1)
        self._file_clear_btn = tk.Button(
            self._file_row,
            text="✕",
            command=self._clear_attached_file,
            bg="#333333",
            fg="white",
            relief="flat",
            padx=4,
            pady=0,
            bd=0,
            cursor="arrow",
        )
        # clear button not packed until a file is attached

        button_frame = tk.Frame(self.root, bg="#111111")
        button_frame.pack(pady=6)

        # Capture button: screenshot + OCR preview
        self.capture_button = tk.Button(
            button_frame,
            text="Capture",
            command=self.capture_screen_and_answer,
            width=12,
            bg="#e69900",
            fg="black",
            relief="flat"
        )
        self.capture_button.pack(side="left", padx=6)

        close_button = tk.Button(
            button_frame,
            text="Exit",
            command=self.stop,
            width=12,
            bg="#666666",
            fg="white",
            relief="flat"
        )
        close_button.pack(side="left", padx=6)

        # allow moving the window by dragging the small header (or the grip)
        header_frame.bind("<ButtonPress-1>", self._start_move)
        header_frame.bind("<B1-Motion>", self._on_move)
        grip.bind("<ButtonPress-1>", self._start_move)
        grip.bind("<B1-Motion>", self._on_move)
        self.root.bind("<Escape>", lambda event: self.stop())
        # global keyboard shortcut for quick full-screen capture
        try:
            self.root.bind_all('<Control-Shift-s>', lambda e: self.capture_screen_and_answer())
        except Exception:
            pass

        # capture state
        self._last_ocr = None
        self._capture_cancelled = False
        self._pending_auto_send_id = None
        self._response_stop_requested = False

        # file attachment state
        self._attached_file = None

        # drag-and-drop: register widgets as drop targets when tkinterdnd2 is available
        if getattr(self, '_dnd_available', False):
            try:
                for widget in (self.log, self.prompt_entry, self._file_row, self._file_indicator_label):
                    try:
                        widget.drop_target_register(_tkdnd.DND_FILES)
                        widget.dnd_bind('<<Drop>>', self._handle_drop)
                    except Exception:
                        pass
            except Exception:
                pass
        else:
            # tkinterdnd2 unavailable — use Win32 WM_DROPFILES fallback (GIL-safe)
            self.root.after(350, self._setup_win32_drag_drop)

        # Ctrl+V on prompt_entry and log: attach file if clipboard holds one, else normal paste
        self.prompt_entry.bind('<Control-v>', self._on_paste_intercept)
        self.log.bind('<Control-v>', self._on_paste_intercept)
        # Ctrl+V on the drop-zone row too
        self._file_row.bind('<Control-v>', self._on_paste_intercept)
        self._file_indicator_label.bind('<Control-v>', self._on_paste_intercept)
        self._file_row.bind('<Button-1>', lambda e: self._file_row.focus_set())
        self._file_indicator_label.bind('<Button-1>', lambda e: self._file_row.focus_set())

    def _set_selected_model(self, model_name: str):
        self.selected_model = model_name
        self._refresh_model_buttons()

    def _refresh_model_buttons(self):
        active_bg = "#44b94e"
        active_fg = "white"
        inactive_bg = "#333333"
        inactive_fg = "#dddddd"

        if hasattr(self, "gpt4o_button"):
            active = self.selected_model == "gpt-4.1"
            self.gpt4o_button.config(bg=active_bg if active else inactive_bg, fg=active_fg if active else inactive_fg)

        if hasattr(self, "gpt5_button"):
            active = self.selected_model == "gpt-5"
            self.gpt5_button.config(bg=active_bg if active else inactive_bg, fg=active_fg if active else inactive_fg)

    def _has_active_request(self) -> bool:
        future = getattr(self, '_current_ai_future', None)
        return bool(future and not future.done())

    def _set_stop_button_enabled(self, enabled: bool):
        try:
            self.stop_button.config(state="normal" if enabled else "disabled")
        except Exception:
            pass

    def _register_active_future(self, future):
        self._current_ai_future = future
        self._response_stop_requested = False
        self._set_stop_button_enabled(True)

    def _clear_active_future(self, future=None):
        current = getattr(self, '_current_ai_future', None)
        if future is not None and current is not future:
            return
        try:
            delattr(self, '_current_ai_future')
        except Exception:
            pass
        self._response_stop_requested = False
        if not self._pending_auto_send_id:
            self._set_stop_button_enabled(False)

    def _is_cancelled_request_error(self, future, exc: Exception) -> bool:
        return bool(
            getattr(self, '_response_stop_requested', False)
            or (future and future.cancelled())
            or isinstance(exc, (asyncio.CancelledError, concurrent.futures.CancelledError))
        )

    def _handle_stop_action(self):
        if self._pending_auto_send_id:
            self._cancel_pending_capture()
            return
        self._stop_active_request()

    def _stop_active_request(self):
        if not self._has_active_request():
            self._set_stop_button_enabled(False)
            return
        self._response_stop_requested = True
        self._set_stop_button_enabled(False)
        self._cancel_current_stream(remove_partial=False)
        try:
            self.status_label.config(text="Response stopped.", fg="#ffb347")
        except Exception:
            pass

    def _on_move(self, event):
        x = self.root.winfo_x() + event.x - self.drag_offset[0]
        y = self.root.winfo_y() + event.y - self.drag_offset[1]
        self.root.geometry(f"+{x}+{y}")

    def _start_move(self, event):
        """Begin window move: record pointer offset within the widget."""
        try:
            # event.x/event.y are widget-local coordinates at press
            self.drag_offset = (event.x, event.y)
        except Exception:
            self.drag_offset = (0, 0)

    def copy_selection(self, event=None):
        try:
            selected = self.log.selection_get()
            self.root.clipboard_clear()
            self.root.clipboard_append(selected)
        except tk.TclError:
            pass
        return "break"

    def select_all(self, event=None):
        self.log.focus_set()
        self.log.tag_add(tk.SEL, "1.0", tk.END)
        self.log.mark_set(tk.INSERT, "1.0")
        self.log.see(tk.INSERT)
        return "break"

    # --- chunk coalescing: stream tokens land here from the async thread ---
    # Instead of one root.after(0,...) per token (~8-15ms Windows overhead each),
    # tokens are pushed into a deque and a single 16ms timer drains them all at once.

    def _start_chunk_flush_loop(self):
        """Start the 16ms repeating flush timer if not already running."""
        if not getattr(self, '_flush_loop_running', False):
            self._flush_loop_running = True
            self.root.after(16, self._flush_chunk_queue)

    def _flush_chunk_queue(self):
        """Drain pending chunks and write them to the log in one shot."""
        try:
            q = getattr(self, '_chunk_queue', None)
            if q is None:
                self._flush_loop_running = False
                return

            chunks = []
            while q:
                chunks.append(q.popleft())

            if chunks:
                combined = "".join(chunks)
                try:
                    pos = getattr(self, '_thinking_placeholder_start', None)
                    if pos is not None:
                        self.log.delete(pos, tk.END)
                        del self._thinking_placeholder_start
                except Exception:
                    pass
                if not hasattr(self, '_assistant_buffer'):
                    self._assistant_buffer = ''
                self._assistant_buffer += combined
                self.log.insert(tk.END, combined, "assistant")
                if not getattr(self, '_user_scrolled', False):
                    self.log.see(tk.END)

            # keep the loop alive while a stream is active
            if getattr(self, '_chunk_queue', None) is not None:
                self.root.after(16, self._flush_chunk_queue)
            else:
                self._flush_loop_running = False
        except Exception:
            self._flush_loop_running = False

    def _append_assistant_chunk(self, text: str):
        """Called from async thread — push token into queue (thread-safe deque append)."""
        try:
            self._chunk_queue.append(text)
        except AttributeError:
            pass

    def _begin_assistant_stream(self):
        self._chunk_queue = collections.deque()
        self._assistant_stream_anchor = self.log.index(tk.END)
        self._assistant_buffer = ''
        self._user_scrolled = False
        self.log.insert(tk.END, "Assistant: ", "assistant")
        try:
            self.log.see(tk.END)
        except Exception:
            pass
        self._start_chunk_flush_loop()

    def _cancel_current_stream(self, remove_partial: bool = True):
        """Attempt to cancel any in-progress AI future and optionally remove partial assistant output."""
        try:
            fut = getattr(self, '_current_ai_future', None)
            if fut and not fut.done():
                try:
                    fut.cancel()
                except Exception:
                    pass
        except Exception:
            pass

        if remove_partial:
            # remove partial assistant text from the UI if a stream anchor exists
            try:
                anchor = getattr(self, '_assistant_stream_anchor', None)
                if anchor:
                    def _del():
                        try:
                            self.log.delete(anchor, tk.END)
                        except Exception:
                            pass
                    self.root.after(0, _del)
                    try:
                        delattr(self, '_assistant_stream_anchor')
                    except Exception:
                        pass
            except Exception:
                pass

    def _finish_assistant_stream(self):
        # drain any remaining queued chunks before closing
        try:
            q = getattr(self, '_chunk_queue', None)
            if q:
                combined = "".join(q)
                q.clear()
                if combined:
                    if not hasattr(self, '_assistant_buffer'):
                        self._assistant_buffer = ''
                    self._assistant_buffer += combined
                    self.log.insert(tk.END, combined, "assistant")
        except Exception:
            pass
        # stop the flush loop
        try:
            self._chunk_queue = None
        except Exception:
            pass
        try:
            pos = getattr(self, '_thinking_placeholder_start', None)
            if pos is not None:
                self.log.delete(pos, tk.END)
                del self._thinking_placeholder_start
        except Exception:
            pass
        try:
            content = getattr(self, '_assistant_buffer', '')
            self._last_assistant = content

            if content:
                self.log.insert(tk.END, "\n", "assistant")
                if not getattr(self, '_user_scrolled', False):
                    self.log.see(tk.END)
            else:
                anchor = getattr(self, '_assistant_stream_anchor', None)
                if anchor:
                    self.log.delete(anchor, tk.END)

            if hasattr(self, '_assistant_buffer'):
                delattr(self, '_assistant_buffer')
            if hasattr(self, '_assistant_stream_anchor'):
                delattr(self, '_assistant_stream_anchor')
        except Exception:
            self._last_assistant = None

    def _hydrate_stream_buffer_from_result(self, response: str | None):
        if not response:
            return
        try:
            existing = getattr(self, '_assistant_buffer', '')
            if existing and existing.strip():
                return
            self._assistant_buffer = response
        except Exception:
            pass

    def _show_context_menu(self, event):
        try:
            self.text_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.text_menu.grab_release()

    def _on_user_scroll(self, event=None):
        # mark that the user interacted with the text area to stop auto-anchoring
        try:
            self._user_scrolled = True
        except Exception:
            self._user_scrolled = True
        return None

    def _on_lang_change(self, value):
        # called when user selects a different language from the dropdown
        try:
            target = value
        except Exception:
            target = self.lang_var.get()
        # cancel any in-progress stream and if there's a recent assistant response, auto-convert
        try:
            self._cancel_current_stream()
        except Exception:
            pass
        # debounce rapid language changes: schedule conversion after short delay
        try:
            if getattr(self, '_pending_lang_job', None):
                try:
                    self.root.after_cancel(self._pending_lang_job)
                except Exception:
                    pass
        except Exception:
            pass

        def _schedule():
            last = getattr(self, '_last_assistant', None)
            if last and last.strip():
                self.log_message('system', f'Converting last assistant response to {target}...')
                threading.Thread(target=self._convert_last_response_to, args=(target,), daemon=True).start()
            else:
                self.log_message('system', f'Language set to {target}. No previous assistant response to convert.')

        try:
            self._pending_lang_job = self.root.after(300, _schedule)
        except Exception:
            # fallback: run immediately
            _schedule()

    # ----------------- Interview listener callbacks -----------------
    def _on_listener_transcript(self, transcript: str):
        try:
            def _show():
                # Insert user message directly (not via log_message) so we can
                # record the exact position right after it without a scheduling gap.
                self.log.insert(tk.END, f"You: {transcript}\n", "user")
                self.log.see(tk.END)
            self.root.after(0, _show)
        except Exception:
            pass

    def _on_listener_answer_delta(self, token: str):
        """Called for each streamed token from the listener GPT call."""
        def _append():
            try:
                if not hasattr(self, '_listener_buffer'):
                    self._listener_buffer = ''
                self._listener_buffer += token
                self.log.insert(tk.END, token, "assistant")
                if not getattr(self, '_user_scrolled', False):
                    self.log.see(tk.END)
            except Exception:
                pass
        self.root.after(0, _append)

    def _on_listener_answer(self, full_reply: str):
        """Called at end of stream with the assembled full reply (for _last_assistant)."""
        def _finish():
            # Always clear Thinking... first — defensive regardless of streaming state
            try:
                pos = getattr(self, '_listener_thinking_start', None)
                if pos is not None:
                    self.log.delete(pos, tk.END)
                    del self._listener_thinking_start
            except Exception:
                pass

            try:
                reply = full_reply or getattr(self, '_listener_buffer', '') or ''
                if reply:
                    self._last_assistant = reply
                    self.log.insert(tk.END, '\n', 'assistant')
                    if not getattr(self, '_user_scrolled', False):
                        self.log.see(tk.END)
                elif hasattr(self, '_listener_stream_anchor'):
                    self.log.delete(self._listener_stream_anchor, tk.END)
                if hasattr(self, '_listener_stream_anchor'):
                    del self._listener_stream_anchor
                if hasattr(self, '_listener_buffer'):
                    del self._listener_buffer
            except Exception:
                pass
        self.root.after(0, _finish)

    def _toggle_listener(self):
        # Start or stop the interview listener in a non-disruptive way
        try:
            if LiveInterviewListener is None:
                self.log_message('system', 'Live listener not available (missing dependency).')
                return

            if not self._listener or not getattr(self._listener, '_running', False):
                try:
                    self._listener = LiveInterviewListener(
                        self.ai, self.loop,
                        on_transcript=self._on_listener_transcript,
                        on_answer=self._on_listener_answer,
                        on_status=self._on_listener_status,
                        preferred_device=getattr(self, '_selected_device', None),
                        selected_model=self.selected_model,
                    )
                    self._listener.on_answer_delta = self._on_listener_answer_delta
                    self._listener.start()
                    try:
                        self.listener_button.config(bg="#ff4444")
                    except Exception:
                        pass
                    self.log_message('system', 'Interview listener started.')
                except Exception as exc:
                    self.log_message('system', f'Listener start failed: {exc}')
            else:
                try:
                    self._listener.stop()
                except Exception:
                    pass
                try:
                    self.listener_button.config(bg="#333333")
                except Exception:
                    pass
                self.log_message('system', 'Interview listener stopped.')
        except Exception as exc:
            try:
                self.log_message('system', f'Listener toggle error: {exc}')
            except Exception:
                pass

    def _on_listener_status(self, status: str):
        # receive lifecycle and error updates from LiveInterviewListener
        # noisy operational messages update a status label only, not the log
        _NOISY = ('rms ', 'flushing ', 'force-flush', 'partial-mark', 'filtered:', 'stream open')
        try:
            if status == 'started':
                self.root.after(0, lambda: self.log_message('system', 'Interview listener thread started'))
            elif status == 'stopped':
                self.root.after(0, lambda: self.log_message('system', 'Interview listener thread stopped'))
            elif status == 'answer:streaming_start':
                def _begin_stream():
                    try:
                        pos = getattr(self, '_listener_thinking_start', None)
                        if pos is not None:
                            self.log.delete(pos, tk.END)
                            del self._listener_thinking_start
                    except Exception:
                        pass
                    self._listener_stream_anchor = self.log.index(tk.END)
                    self._listener_buffer = ''
                    self.log.insert(tk.END, 'Assistant: ', 'assistant')
                    if not getattr(self, '_user_scrolled', False):
                        self.log.see(tk.END)
                self.root.after(0, _begin_stream)
            elif status == 'answer:streaming_end':
                pass  # handled in _on_listener_answer
            elif status == 'answer:skipped':
                # Filter blocked the GPT call — clear any Thinking... placeholder
                def _clear_thinking():
                    try:
                        pos = getattr(self, '_listener_thinking_start', None)
                        if pos is not None:
                            self.log.delete(pos, tk.END)
                            del self._listener_thinking_start
                    except Exception:
                        pass
                self.root.after(0, _clear_thinking)
            elif status == 'answer:final':
                pass  # non-streaming path — answer is delivered via on_answer
            elif status and (status.startswith('error:') or status.startswith('transcribe_error') or status.startswith('stream_open_error')):
                self.root.after(0, lambda s=status: self.log_message('system', f'Listener error: {s}'))
            elif status and any(status.startswith(p) for p in _NOISY):
                # update compact status label without touching the log scroll
                self.root.after(0, lambda s=status: self._update_listener_status_label(s))
            else:
                self.root.after(0, lambda s=status: self.log_message('system', f'Listener: {s}'))
        except Exception:
            pass

    def _update_listener_status_label(self, msg: str):
        """Update mic status bar in-place without scrolling the log."""
        try:
            lbl = getattr(self, '_listener_status_lbl', None)
            if lbl:
                lbl.config(text=msg[:80])
        except Exception:
            pass

    def _convert_last_response_to(self, target_lang: str):
        try:
            text = getattr(self, '_last_assistant', '')
            if not text or not text.strip():
                return
            if self._has_active_request():
                self.root.after(0, lambda: self.log_message('system', 'A request is already running.'))
                return
            # build a conversion prompt that asks for language translation and preserving/adding human comments
            conv = (
                f"Convert the following code and explanation into {target_lang}. Preserve the behavior exactly. "
                "If the original code contains comments, convert them to the target language's single-line comment syntax and keep them natural and human-sounding. "
                "If the original code lacks line comments, add short human-sounding comments to every line explaining what it does. "
                "Use the target language's idiomatic style and provide runnable code blocks where applicable.\n\n"
                "Original assistant response:\n" + text
            )
            # Ask converter to include an example usage snippet when producing code
            conv += ("\n\nAdditionally: when you output code, include at least two short example usages showing how to call or run the converted code with concrete example values. Place these examples immediately after the code block and before the explanation. "
                     "After the examples, provide a concise explanation written for an interviewer: describe why you chose this approach, trade-offs, complexity, and any edge-cases handled.")
            # stream the converted result into the UI
            selected_model = self.selected_model
            self.root.after(0, self._begin_assistant_stream)
            future = asyncio.run_coroutine_threadsafe(
                self.ai.ask_gpt(conv, mode="chat", stream=True, on_delta=self._append_assistant_chunk, selected_model=selected_model, include_context=False),
                self.loop,
            )
            self._register_active_future(future)
            try:
                future.result(timeout=300)
                self.root.after(0, self._finish_assistant_stream)
            except Exception as exc:
                self.root.after(0, self._finish_assistant_stream)
                if not self._is_cancelled_request_error(future, exc):
                    self.root.after(0, lambda exc=exc: self.log_message('assistant', f'Conversion error: {type(exc).__name__}: {exc}'))
            finally:
                self._clear_active_future(future)
        except Exception as exc:
            try:
                self.root.after(0, lambda exc=exc: self.log_message('assistant', f'Conversion unexpected error: {type(exc).__name__}: {exc}'))
            except Exception:
                pass

    def _make_window_layered(self, hwnd):
        GWL_EXSTYLE = -20
        WS_EX_LAYERED = 0x80000
        exstyle = ctypes.windll.user32.GetWindowLongW(hwnd, GWL_EXSTYLE)
        if exstyle & WS_EX_LAYERED:
            return
        ctypes.windll.user32.SetWindowLongW(hwnd, GWL_EXSTYLE, exstyle | WS_EX_LAYERED)

    def _set_display_affinity(self, hwnd, affinity):
        fn = ctypes.windll.user32.SetWindowDisplayAffinity
        fn.argtypes = [wintypes.HWND, wintypes.DWORD]
        fn.restype = wintypes.BOOL
        return fn(hwnd, affinity)

    def _apply_taskbar_hide(self):
        """Remove the app from the Windows taskbar by setting WS_EX_TOOLWINDOW and clearing WS_EX_APPWINDOW.

        This runs shortly after UI build to ensure the window exists and is hidden from taskbar/Alt-Tab when packaged.
        """
        if platform.system() != 'Windows':
            return
        try:
            GWL_EXSTYLE = -20
            WS_EX_APPWINDOW = 0x00040000
            WS_EX_TOOLWINDOW = 0x00000080
            SWP_NOSIZE = 0x0001
            SWP_NOMOVE = 0x0002
            SWP_NOZORDER = 0x0004
            SWP_FRAMECHANGED = 0x0020

            self.root.update_idletasks()
            hwnd = self.root.winfo_id()
            # ensure we operate on the top-level window
            try:
                hwnd = ctypes.windll.user32.GetAncestor(hwnd, 2)
            except Exception:
                pass

            ex = ctypes.windll.user32.GetWindowLongW(hwnd, GWL_EXSTYLE)
            # clear APPWINDOW and set TOOLWINDOW
            ex = (ex & ~WS_EX_APPWINDOW) | WS_EX_TOOLWINDOW
            ctypes.windll.user32.SetWindowLongW(hwnd, GWL_EXSTYLE, ex)
            # refresh window frame so changes take effect
            ctypes.windll.user32.SetWindowPos(hwnd, 0, 0, 0, 0, 0, SWP_NOMOVE | SWP_NOSIZE | SWP_NOZORDER | SWP_FRAMECHANGED)
        except Exception:
            pass

    def _hide_from_capture(self):
        if platform.system() != "Windows":
            return

        winver = sys.getwindowsversion()
        if winver.major < 10 or (winver.major == 10 and winver.build < 14393):
            if not self._capture_error_logged:
                self.log_message("system", "Capture protection unavailable: Windows version does not support screen-capture exclusion.")
                self._capture_error_logged = True
            return

        try:
            self.root.update_idletasks()
            hwnd = self.root.winfo_id()
            hwnd = ctypes.windll.user32.GetAncestor(hwnd, 2)
            self._make_window_layered(hwnd)

            if self._capture_affinity is None:
                for affinity in (WDA_EXCLUDEFROMCAPTURE, WDA_MONITOR):
                    result = self._set_display_affinity(hwnd, affinity)
                    if result:
                        self._capture_affinity = affinity
                        break
                if self._capture_affinity is None:
                    raise ctypes.WinError()
            else:
                result = self._set_display_affinity(hwnd, self._capture_affinity)
                if not result:
                    raise ctypes.WinError()
        except Exception as exc:
            if not self._capture_error_logged:
                self.log_message("system", f"Capture protection unavailable: {exc}")
                self._capture_error_logged = True
        finally:
            self.root.after(2000, self._hide_from_capture)

    def start_recording(self):
        # audio recording disabled in lightweight UI
        self.log_message('system', 'Start recording disabled in lightweight capture mode.')

    def stop_recording(self):
        # audio recording disabled
        self.log_message('system', 'Stop recording disabled in lightweight capture mode.')

    def stop(self):
        try:
            self._cancel_current_stream(remove_partial=False)
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass

    def _capture_loop(self):
        # audio capture loop removed in lightweight UI
        return

    def _browse_file(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Attach a file",
            filetypes=[
                ("All supported", "*.csv *.xlsx *.xls *.log *.txt *.json *.py *.js *.ts *.md *.xml *.yaml *.yml *.html *.htm *.sql"),
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx *.xls"),
                ("Log / Text", "*.log *.txt"),
                ("All files", "*.*"),
            ]
        )
        if path:
            self._attach_file(path)

    def _attach_file(self, path: str):
        self._attached_file = path
        name = Path(path).name
        try:
            self._file_indicator_label.config(
                text=f"📎  {name}",
                fg="#8be9fd",
                bg="#1e1e2e",
            )
            self._file_row.config(bg="#1e1e2e")
            self._file_clear_btn.pack(side="right", padx=2, pady=1)
        except Exception:
            pass

    def _clear_attached_file(self):
        self._attached_file = None
        try:
            self._file_clear_btn.pack_forget()
            self._file_indicator_label.config(
                text="  Drop file here  ·  Ctrl+V  ·  📎  ",
                fg="#555577",
                bg="#1a1a2e",
            )
            self._file_row.config(bg="#1a1a2e")
        except Exception:
            pass

    def _try_paste_file_from_clipboard(self) -> bool:
        """Check Windows clipboard for CF_HDROP (files copied in Explorer).
        Returns True and attaches the first file if found."""
        import ctypes
        from ctypes import wintypes
        CF_HDROP = 15
        try:
            user32 = ctypes.windll.user32
            shell32 = ctypes.windll.shell32

            user32.IsClipboardFormatAvailable.restype = wintypes.BOOL
            if not user32.IsClipboardFormatAvailable(CF_HDROP):
                return False

            user32.OpenClipboard.restype = wintypes.BOOL
            if not user32.OpenClipboard(None):
                return False

            path = None
            try:
                user32.GetClipboardData.restype = ctypes.c_void_p
                h = user32.GetClipboardData(CF_HDROP)
                if h:
                    shell32.DragQueryFileW.restype = ctypes.c_uint
                    shell32.DragQueryFileW.argtypes = [
                        ctypes.c_void_p, ctypes.c_uint,
                        ctypes.c_wchar_p, ctypes.c_uint,
                    ]
                    n = shell32.DragQueryFileW(h, 0, None, 0) + 1
                    buf = ctypes.create_unicode_buffer(n)
                    shell32.DragQueryFileW(h, 0, buf, n)
                    path = buf.value
            finally:
                user32.CloseClipboard()

            if path and Path(path).exists():
                self._attach_file(path)
                return True
        except Exception:
            pass
        return False

    def _on_paste_intercept(self, event=None):
        """Intercept Ctrl+V: attach file if clipboard has CF_HDROP, else allow normal paste."""
        if self._try_paste_file_from_clipboard():
            return "break"
        return None  # propagate to widget's own paste handler

    def _read_file_content(self, path: str) -> str:
        ext = Path(path).suffix.lower()
        max_chars = 40000
        try:
            if ext == ".csv":
                rows = []
                with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
                    reader = csv.reader(f)
                    for i, row in enumerate(reader):
                        rows.append(", ".join(row))
                        if i >= 500:
                            rows.append("... (truncated at 500 rows)")
                            break
                return "\n".join(rows)[:max_chars]

            if ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    lines = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        lines.append(f"[Sheet: {sheet_name}]")
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            lines.append(", ".join("" if v is None else str(v) for v in row))
                            if i >= 500:
                                lines.append("... (truncated)")
                                break
                    return "\n".join(lines)[:max_chars]
                except ImportError:
                    pass
                try:
                    import xlrd
                    wb = xlrd.open_workbook(path)
                    lines = []
                    for sheet in wb.sheets():
                        lines.append(f"[Sheet: {sheet.name}]")
                        for i in range(min(sheet.nrows, 501)):
                            if i == 500:
                                lines.append("... (truncated)")
                                break
                            lines.append(", ".join(str(v) for v in sheet.row_values(i)))
                    return "\n".join(lines)[:max_chars]
                except ImportError:
                    pass
                return "[Excel support requires openpyxl or xlrd: pip install openpyxl]"

            # text-based files (log, txt, py, js, json, md, sql, etc.)
            for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
                try:
                    with open(path, encoding=enc, errors="strict") as f:
                        return f.read(max_chars)
                except (UnicodeDecodeError, LookupError):
                    continue
            with open(path, encoding="utf-8", errors="replace") as f:
                return f.read(max_chars)
        except Exception as exc:
            return f"[Error reading file: {exc}]"

    def _handle_drop(self, event):
        try:
            raw = (event.data or "").strip()
            # tkinterdnd2 wraps paths containing spaces in {} on Windows
            import re as _re
            braced = _re.findall(r'\{([^}]+)\}', raw)
            if braced:
                path = braced[0].strip()
            else:
                # strip quotes and surrounding whitespace
                path = raw.strip('"').strip("'").strip()
            if path and Path(path).exists():
                self._attach_file(path)
            elif path:
                # path may be valid but relative; try as-is
                self._attach_file(path)
        except Exception:
            pass

    def _setup_win32_drag_drop(self):
        """GIL-safe WM_DROPFILES fallback (used only when tkinterdnd2 is unavailable).

        The WINFUNCTYPE wndproc ONLY calls pure C APIs — no Python object creation,
        no self.root.after() — to avoid PyEval_RestoreThread(NULL) crashes in Python 3.14.
        A daemon watcher thread waits on a Win32 event and calls after() safely.
        """
        import platform
        if platform.system() != 'Windows':
            return
        import ctypes
        import threading
        from ctypes import wintypes
        WM_DROPFILES = 0x0233
        GWLP_WNDPROC = -4
        try:
            self.root.update_idletasks()
            user32 = ctypes.windll.user32
            shell32 = ctypes.windll.shell32
            kernel32 = ctypes.windll.kernel32

            hwnd = user32.GetAncestor(self.root.winfo_id(), 2)
            shell32.DragAcceptFiles(hwnd, True)

            # Pre-allocated C buffer — written by C code, read by Python watcher thread
            _buf = (ctypes.c_wchar * 4096)()
            # Manual-reset Win32 event for cross-thread signalling
            _event = kernel32.CreateEventW(None, True, False, None)

            shell32.DragQueryFileW.restype = ctypes.c_uint
            shell32.DragQueryFileW.argtypes = [
                ctypes.c_void_p, ctypes.c_uint, ctypes.c_wchar_p, ctypes.c_uint,
            ]
            shell32.DragFinish.argtypes = [ctypes.c_void_p]

            _LRESULT = ctypes.c_ssize_t
            _WNDPROC = ctypes.WINFUNCTYPE(
                _LRESULT, wintypes.HWND, wintypes.UINT, wintypes.WPARAM, wintypes.LPARAM,
            )

            user32.GetWindowLongPtrW.restype = _LRESULT
            _orig = user32.GetWindowLongPtrW(hwnd, GWLP_WNDPROC)
            user32.CallWindowProcW.restype = _LRESULT
            user32.CallWindowProcW.argtypes = [
                _LRESULT, wintypes.HWND, wintypes.UINT, wintypes.WPARAM, wintypes.LPARAM,
            ]

            def _wndproc(h, msg, wp, lp):
                if msg == WM_DROPFILES:
                    # ONLY pure C API calls here — no Python objects, no after().
                    # Avoids PyEval_RestoreThread(NULL) crash when GIL was entered
                    # via PyGILState_Ensure (callback context, not main-thread context).
                    shell32.DragQueryFileW(wp, 0, _buf, 4096)
                    shell32.DragFinish(wp)
                    kernel32.SetEvent(_event)
                    return 0
                return user32.CallWindowProcW(_orig, h, msg, wp, lp)

            self._drop_wndproc = _WNDPROC(_wndproc)  # must stay alive
            user32.SetWindowLongPtrW.restype = _LRESULT
            user32.SetWindowLongPtrW.argtypes = [wintypes.HWND, ctypes.c_int, _LRESULT]
            user32.SetWindowLongPtrW(
                hwnd, GWLP_WNDPROC,
                ctypes.cast(self._drop_wndproc, ctypes.c_void_p).value,
            )

            # Watcher thread: proper Python thread state — safe to call after() here
            def _watcher():
                WAIT_OBJECT_0 = 0
                INFINITE = 0xFFFFFFFF
                while True:
                    r = kernel32.WaitForSingleObject(_event, INFINITE)
                    if r != WAIT_OBJECT_0:
                        break  # handle closed or wait failed — exit cleanly
                    path = _buf.value
                    kernel32.ResetEvent(_event)
                    if path:
                        self.root.after(0, lambda p=path: self._attach_file(p))

            threading.Thread(target=_watcher, daemon=True).start()
        except Exception:
            pass

    def send_prompt(self, event=None):
        prompt = self.prompt_entry.get().strip()
        file_path = getattr(self, '_attached_file', None)
        if not prompt and not file_path:
            return
        if self._has_active_request():
            self.log_message("system", "A request is already running.")
            return

        self.prompt_entry.delete(0, tk.END)
        self._clear_attached_file()
        display = prompt if prompt else f"[File: {Path(file_path).name}]"
        self.log_message("user", display)
        self.send_button.config(state="disabled")

        threading.Thread(target=self._send_prompt_thread, args=(prompt, self.selected_model, file_path), daemon=True).start()

    def _send_prompt_thread(self, prompt: str, selected_model: str, file_path: str | None = None):
        try:
            combined = prompt
            if file_path:
                file_content = self._read_file_content(file_path)
                file_name = Path(file_path).name
                if file_content and not file_content.startswith("[Error"):
                    file_section = f"\n\n--- File: {file_name} ---\n{file_content}\n--- End of file ---"
                    combined = (
                        (prompt + file_section) if prompt
                        else (
                            f"Analyze the following file ({file_name}) and provide relevant insights, "
                            f"code, or answers based on its content:{file_section}"
                        )
                    )
                else:
                    err = file_content or "unreadable"
                    self.root.after(0, lambda e=err: self.log_message('system', f'File read error: {e}'))
                    if not prompt:
                        self.root.after(0, lambda: self.send_button.config(state="normal"))
                        self.root.after(0, lambda: self.status_label.config(text="Ready for the next prompt.", fg="#50fa7b"))
                        return

            self.root.after(0, lambda: self.status_label.config(text="Answering now...", fg="#a6e22e"))
            self.root.after(0, self._begin_assistant_stream)

            future = asyncio.run_coroutine_threadsafe(
                self.ai.ask_gpt(combined, mode="chat", stream=True, on_delta=self._append_assistant_chunk, selected_model=selected_model, include_context=True),
                self.loop,
            )
            self._register_active_future(future)
            try:
                # extended timeout to avoid truncated/early termination of streaming responses
                response = future.result(timeout=300)
                self.root.after(0, lambda response=response: self._hydrate_stream_buffer_from_result(response))
                self.root.after(0, self._finish_assistant_stream)
            except Exception as exc:
                self.root.after(0, self._finish_assistant_stream)
                if not self._is_cancelled_request_error(future, exc):
                    self.root.after(0, lambda exc=exc: self.log_message('assistant', f'Chat error: {type(exc).__name__}: {exc}'))
            self._clear_active_future(future)
        except Exception as exc:
            try:
                self.root.after(0, lambda exc=exc: self.log_message('assistant', f'Chat error: {type(exc).__name__}: {exc}'))
            except Exception:
                pass
        finally:
            self.root.after(0, lambda: self.send_button.config(state="normal"))
            try:
                self.root.after(0, lambda: self.status_label.config(text="Ready for the next prompt.", fg="#50fa7b"))
            except Exception:
                pass

    def run(self):
        """Run the Tkinter main loop for the overlay."""
        self.root.mainloop()

    def _configure_tesseract(self, pytesseract) -> bool:
        """Ensure pytesseract points to a valid tesseract executable.

        Returns True if configured, False otherwise.
        """
        # prefer explicit config, then environment, then common install path
        tpath = None
        try:
            from config import TESSERACT_CMD
            tpath = TESSERACT_CMD
        except Exception:
            tpath = None

        if not tpath:
            tpath = os.environ.get('TESSERACT_CMD') or os.environ.get('TESSERACT_PATH')

        if tpath and os.path.exists(tpath):
            pytesseract.pytesseract.tesseract_cmd = tpath
            return True

        # try default Windows install path
        default = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(default):
            pytesseract.pytesseract.tesseract_cmd = default
            return True

        return False

    def _get_overlay_hwnd(self):
        if platform.system() != 'Windows':
            return None
        try:
            hwnd = self.root.winfo_id()
            return ctypes.windll.user32.GetAncestor(hwnd, 2)
        except Exception:
            return None

    def _get_foreground_window_bbox(self):
        if platform.system() != 'Windows':
            return None
        try:
            user32 = ctypes.windll.user32
            hwnd = user32.GetForegroundWindow()
            if not hwnd:
                return None
            overlay_hwnd = self._get_overlay_hwnd()
            if overlay_hwnd and hwnd == overlay_hwnd:
                return None

            rect = wintypes.RECT()
            if not user32.GetWindowRect(hwnd, ctypes.byref(rect)):
                return None

            left, top, right, bottom = rect.left, rect.top, rect.right, rect.bottom
            if right - left < 120 or bottom - top < 120:
                return None
            return (left, top, right, bottom)
        except Exception:
            return None

    def _grab_preferred_capture_image(self, image_grab_module):
        bbox = self._get_foreground_window_bbox()
        if bbox:
            try:
                return image_grab_module.grab(bbox=bbox), 'active-window'
            except Exception:
                pass
        return image_grab_module.grab(all_screens=True), 'all-screens'

    def _ocr_confidence(self, image, pytesseract) -> float:
        try:
            data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
        except Exception:
            return -1.0

        confidences = []
        for raw_conf, raw_text in zip(data.get('conf', []), data.get('text', [])):
            text = (raw_text or '').strip()
            if not text:
                continue
            try:
                conf = float(raw_conf)
            except Exception:
                continue
            if conf >= 0:
                confidences.append(conf)

        if not confidences:
            return -1.0
        return sum(confidences) / len(confidences)

    def capture_screen_and_answer(self):
        """Capture the screen (image), OCR it to text, and send to the AI for an answer.

        This runs the screenshot in a background thread, temporarily hides the overlay
        so it is not captured, and re-shows it after grabbing. It then runs OCR and
        schedules the forced send as usual.
        """

        # perform the screen grab on the main thread (via after) to avoid ImageGrab hanging
        self.log_message('system', 'Starting full-screen capture...')
        try:
            from PIL import ImageGrab
        except Exception:
            self.log_message("system", "Capture failed: Pillow not installed. Install 'pillow' in the venv.")
            return

        # disable capture controls and hide overlay, then grab on the mainloop shortly after
        try:
            self.header_capture.config(state='disabled')
            self.capture_button.config(state='disabled')
        except Exception:
            pass

        def _grab_on_main():
            # withdraw then grab immediately
            try:
                self.root.withdraw()
            except Exception:
                pass

            try:
                img, capture_mode = self._grab_preferred_capture_image(ImageGrab)
                self.log_message('system', f'Capture mode: {capture_mode}')
            except Exception as exc:
                self.log_message('system', f'Screenshot failed: {exc}')
                img = None

            try:
                self.root.deiconify()
            except Exception:
                pass

            try:
                self.header_capture.config(state='normal')
                self.capture_button.config(state='normal')
            except Exception:
                pass

            if img is None:
                return

            # process OCR in background to avoid blocking UI
            threading.Thread(target=self._process_captured_image, args=(img,), daemon=True).start()

        # schedule grab after 150ms to allow withdraw to take effect
        self.root.after(150, _grab_on_main)

    # Region selection capture (faster): user draws a rectangle on screen
    def start_region_selection(self):
        sel = tk.Toplevel(self.root)
        sel.attributes("-fullscreen", True)
        sel.attributes("-alpha", 0.25)
        sel.configure(bg="black")
        sel.overrideredirect(True)

        # Use the toplevel's background color (empty string isn't a valid color)
        canvas = tk.Canvas(sel, cursor="cross", bg=sel.cget('bg'), highlightthickness=0)
        canvas.pack(fill="both", expand=True)

        start = {}
        rect = None

        def on_mouse_down(event):
            start['x'] = event.x_root
            start['y'] = event.y_root

        def on_mouse_move(event):
            nonlocal rect
            # convert screen coordinates to canvas-local coordinates
            x1, y1 = start.get('x', 0), start.get('y', 0)
            x2, y2 = event.x_root, event.y_root
            lx1 = x1 - sel.winfo_rootx()
            ly1 = y1 - sel.winfo_rooty()
            lx2 = x2 - sel.winfo_rootx()
            ly2 = y2 - sel.winfo_rooty()
            canvas.delete('selrect')
            canvas.create_rectangle(lx1, ly1, lx2, ly2, outline='red', width=2, tag='selrect')

        def on_mouse_up(event):
            x1, y1 = start.get('x', 0), start.get('y', 0)
            x2, y2 = event.x_root, event.y_root
            sel.destroy()
            bbox = (min(x1, x2), min(y1, y2), max(x1, x2), max(y1, y2))
            threading.Thread(target=self._capture_region_and_ask, args=(bbox,), daemon=True).start()

        canvas.bind('<ButtonPress-1>', on_mouse_down)
        canvas.bind('<B1-Motion>', on_mouse_move)
        canvas.bind('<ButtonRelease-1>', on_mouse_up)

    def _capture_region_and_ask(self, bbox):
        try:
            from PIL import ImageGrab, Image, ImageOps, ImageFilter
            import pytesseract
        except Exception as exc:
            self.log_message('system', f'Capture failed: missing libs ({exc})')
            return
        try:
            if not self._configure_tesseract(pytesseract):
                self.log_message('system', 'OCR failed: tesseract is not installed or not in PATH. See TESSERACT_INSTALLATION.md')
                return
        except Exception:
            pass

        try:
            img = ImageGrab.grab(bbox=bbox)
        except Exception as exc:
            self.log_message('system', f'Screenshot failed: {exc}')
            return

        try:
            proc = self._preprocess_for_ocr(img)
            text, chosen_psm = self._ocr_best_psm(proc, pytesseract, psm_list=[6,3,11,1])
            self.log_message('system', f'OCR used psm={chosen_psm}')
        except Exception as exc:
            self.log_message('system', f'OCR failed: {exc}')
            return

        if not text or not text.strip():
            self.log_message('system', 'No text detected in selected region.')
            return

        # Log OCR text and schedule forced send
        self.log_message('system', 'Captured OCR text:')
        self.log_message('system', text)
        self._capture_cancelled = False
        try:
            self.cancel_send_button.config(state='normal')
        except Exception:
            pass
        if self._pending_auto_send_id:
            try:
                self.root.after_cancel(self._pending_auto_send_id)
            except Exception:
                pass
        selected_model = self.selected_model
        self._pending_auto_send_id = self.root.after(2000, lambda: threading.Thread(target=self._auto_send_forced, args=(text, selected_model), daemon=True).start())

    def _cancel_pending_capture(self):
        try:
            self._capture_cancelled = True
            if self._pending_auto_send_id:
                try:
                    self.root.after_cancel(self._pending_auto_send_id)
                except Exception:
                    pass
                self._pending_auto_send_id = None
            try:
                self.cancel_send_button.config(state='disabled')
            except Exception:
                pass
            self.log_message('system', 'Capture send cancelled by user.')
        except Exception:
            pass

    def _auto_send_forced(self, text: str, selected_model: str):
        # Runs in background thread
        try:
            if getattr(self, '_capture_cancelled', False):
                return
            if self._has_active_request():
                self.root.after(0, lambda: self.log_message('system', 'A request is already running.'))
                return
            # disable cancel button in UI
            self.root.after(0, lambda: self.cancel_send_button.config(state='disabled'))
            # clear pending id
            self._pending_auto_send_id = None
            lang = getattr(self, 'lang_var', None).get() if getattr(self, 'lang_var', None) else 'Python'
            forced = (
                f"Treat the captured text as the task to solve. Produce a concise, runnable {lang} solution that matches the captured request exactly. "
                "Do not reuse earlier prompts or add unrelated requirements. If the text is ambiguous, state the most practical interpretation briefly and continue. "
                "Prefer the simplest practical real-world implementation. Do not invent extra features such as thread safety, async behavior, rotation, multiple handlers, metadata fields, retries, or scalability unless the captured text asks for them.\n\n"
                "Captured text:\n" + text + (
                    f"\n\nAdditionally: if you include code, keep it concise, runnable, and focused on the requested solution. Respond using {lang} code blocks where applicable. Do not add long explanations unless needed."
                )
            )
            # start streaming
            self.root.after(0, self._begin_assistant_stream)
            future = asyncio.run_coroutine_threadsafe(
                self.ai.ask_gpt(forced, mode="chat", stream=True, on_delta=self._append_assistant_chunk, selected_model=selected_model, include_context=False),
                self.loop,
            )
            self._register_active_future(future)
            try:
                # extended timeout for capture-triggered requests
                response = future.result(timeout=300)
                self.root.after(0, lambda response=response: self._hydrate_stream_buffer_from_result(response))
                self.root.after(0, self._finish_assistant_stream)
            except Exception as exc:
                self.root.after(0, self._finish_assistant_stream)
                if not self._is_cancelled_request_error(future, exc):
                    self.root.after(0, lambda exc=exc: self.log_message('assistant', f'Auto-send error: {type(exc).__name__}: {exc}'))
            finally:
                self._clear_active_future(future)
        except Exception as exc:
            try:
                self.root.after(0, self._finish_assistant_stream)
                self.root.after(0, lambda exc=exc: self.log_message('assistant', f'Auto-send unexpected error: {type(exc).__name__}: {exc}'))
            except Exception:
                pass

    def _capture_ask_thread(self, prompt: str, selected_model: str):
        try:
            if self._has_active_request():
                self.root.after(0, lambda: self.log_message('system', 'A request is already running.'))
                return
            self.root.after(0, self._begin_assistant_stream)
            future = asyncio.run_coroutine_threadsafe(
                self.ai.ask_gpt(prompt, mode="chat", stream=True, on_delta=self._append_assistant_chunk, selected_model=selected_model, include_context=False),
                self.loop,
            )
            self._register_active_future(future)
            try:
                response = future.result(timeout=90)
                self.root.after(0, lambda response=response: self._hydrate_stream_buffer_from_result(response))
                if not response:
                    self.root.after(0, lambda: self.log_message("assistant", "No response received."))
            except Exception as exc:
                if not self._is_cancelled_request_error(future, exc):
                    raise
            finally:
                self.root.after(0, self._finish_assistant_stream)
                self._clear_active_future(future)
        except Exception as exc:
            self.root.after(0, lambda exc=exc: self.log_message("assistant", f"Capture chat error: {type(exc).__name__}: {exc}"))
        finally:
            self.root.after(0, lambda: self.status_label.config(text="Ready for the next prompt.", fg="#50fa7b"))

    def _toggle_listen(self):
        if not self._listening:
            self._start_live_listen()
        else:
            self._stop_live_listen()

    def _test_mic(self):
        import sounddevice as sd
        import numpy as _np
        import wave
        import tempfile
        try:
            tmp = Path(tempfile.gettempdir()) / f"mic_test_{int(time.time()*1000)}.wav"
            sr = getattr(self.recorder, 'samplerate', 16000)
            ch = getattr(self.recorder, 'channels', 1)
            dev = None
            try:
                sel = self.device_var.get()
                dev = int(sel) if sel else None
            except Exception:
                dev = None
            # adapt to device defaults when possible
            try:
                if dev is not None:
                    info = sd.query_devices(dev)
                    sr = int(info.get('default_samplerate', sr))
                    max_in = int(info.get('max_input_channels', ch) or ch)
                    ch = min(ch, max_in) if max_in > 0 else ch
                else:
                    # use default input device
                    dd = sd.default.device
                    idx = None
                    if isinstance(dd, (list, tuple)):
                        idx = dd[0]
                    else:
                        idx = dd
                    try:
                        info = sd.query_devices(idx)
                        sr = int(info.get('default_samplerate', sr))
                        max_in = int(info.get('max_input_channels', ch) or ch)
                        ch = min(ch, max_in) if max_in > 0 else ch
                        dev = idx
                    except Exception:
                        pass
            except Exception:
                pass

            frames = int(sr * 1.0)
            try:
                if dev is not None:
                    rec = sd.rec(frames, samplerate=sr, channels=ch, dtype='int16', device=dev)
                else:
                    rec = sd.rec(frames, samplerate=sr, channels=ch, dtype='int16')
            except Exception as e:
                # fallback: try common rates
                for try_sr in (44100, 48000, 16000):
                    try:
                        sr = try_sr
                        frames = int(sr * 1.0)
                        if dev is not None:
                            rec = sd.rec(frames, samplerate=sr, channels=ch, dtype='int16', device=dev)
                        else:
                            rec = sd.rec(frames, samplerate=sr, channels=ch, dtype='int16')
                        break
                    except Exception:
                        rec = None
                        continue
                if rec is None:
                    raise
            sd.wait()
            arr = _np.frombuffer(rec.tobytes(), dtype=_np.int16)
            rms = float(_np.sqrt(_np.mean((arr.astype('float32') / 32768.0) ** 2))) if arr.size else 0.0
            self.log_message('system', f'Mic test RMS={rms:.4f} (device={dev})')
            try:
                with wave.open(str(tmp), 'wb') as wf:
                    wf.setnchannels(ch)
                    wf.setsampwidth(2)
                    wf.setframerate(sr)
                    wf.writeframes(rec.tobytes())
            except Exception:
                pass
        except Exception as exc:
            self.log_message('system', f'Mic test failed: {exc}')

    def _start_live_listen(self):
        try:
            # start recorder if not running
            self._live_stop_event.clear()
            self._live_last_transcript = ""
            # pick device from menu if present
            try:
                sel = self.device_var.get()
                self._selected_device = int(sel) if sel else None
            except Exception:
                self._selected_device = None

            self._listening = True
            try:
                self.mic_button.config(bg="#ff4444")
            except Exception:
                pass
            self.log_message('system', 'Live listening started...')
            # Start a thread that will probe devices if necessary and then run the live loop
            self._live_thread = threading.Thread(target=self._select_and_start_live, daemon=True)
            self._live_thread.start()
        except Exception as exc:
            self.log_message('system', f'Live listen failed: {exc}')

    def _probe_single_device(self, device_idx, duration=0.6, try_srs=(16000, 44100, 48000)):
        """Record a short snippet from a single device and return (rms, samplerate) or (0.0, None) on failure."""
        try:
            import sounddevice as sd
            import numpy as _np
        except Exception:
            return 0.0, None

        try:
            info = sd.query_devices(int(device_idx))
        except Exception:
            return 0.0, None

        # Build ordered list of sample rates: prefer device default first
        default_sr = None
        try:
            default_sr = int(info.get('default_samplerate'))
        except Exception:
            default_sr = None

        srs = []
        if default_sr:
            srs.append(default_sr)
        for sr in try_srs:
            try:
                if sr in srs:
                    # already included
                    pass
                else:
                    srs.append(sr)
            except Exception:
                continue

        for sr in srs:
            try:
                ch = int(info.get('max_input_channels', 1) or 1)
                frames = int(sr * duration)
                rec = sd.rec(frames, samplerate=sr, channels=min(ch, 1), dtype='int16', device=int(device_idx))
                sd.wait(timeout=int(duration * 1000) + 500)
                import numpy as np
                arr = np.frombuffer(rec.tobytes(), dtype=np.int16)
                if arr.size:
                    rms = float(np.sqrt(np.mean((arr.astype('float32') / 32768.0) ** 2)))
                else:
                    rms = 0.0
                return rms, sr
            except Exception:
                continue
        return 0.0, None

    def _auto_select_working_device(self, duration=0.6, threshold=0.002):
        """Iterate input-capable devices and return first (index, rms) above threshold, or (None, 0.0)."""
        try:
            import sounddevice as sd
            import numpy as _np
        except Exception:
            return None, 0.0

        try:
            devs = sd.query_devices()
        except Exception:
            return None, 0.0

        # gather candidate indices (input-capable)
        candidates = [i for i, d in enumerate(devs) if d.get('max_input_channels', 0) > 0]
        # prefer devices listed in the UI menu first
        try:
            pref = [int(k) for k in getattr(self, '_device_map', {}).keys()]
            for p in reversed(pref):
                if p in candidates:
                    candidates.remove(p)
                    candidates.insert(0, p)
        except Exception:
            pass

        # Probe each candidate and pick the one with the highest RMS
        best_idx = None
        best_rms = 0.0
        for idx in candidates:
            try:
                rms, sr = self._probe_single_device(idx, duration=duration)
                if rms and rms > best_rms:
                    best_rms = rms
                    best_idx = idx
            except Exception:
                continue

        # Accept any device with a small but non-zero RMS (handles low-gain mics)
        min_accept = 1e-5
        if best_idx is not None and best_rms >= min_accept:
            return best_idx, best_rms

        return None, 0.0

    def _select_and_start_live(self):
        """Background thread entry: ensure a working input device is selected, then run the live loop."""
        # prefer explicit UI selection
        try:
            sel = getattr(self, '_selected_device', None)
        except Exception:
            sel = None

        chosen = None
        if sel is not None:
            try:
                rms, sr = self._probe_single_device(sel)
                if rms and rms > 0.0015:
                    chosen = sel
            except Exception:
                chosen = None

        if chosen is None:
            idx, rms = self._auto_select_working_device()
            if idx is not None:
                chosen = idx
                try:
                    self.root.after(0, lambda: self.log_message('system', f'Auto-selected device {idx} (rms={rms:.4f})'))
                except Exception:
                    pass
            else:
                try:
                    self.root.after(0, lambda: self.log_message('system', 'No working input device found. Check microphone permissions and device selection.'))
                except Exception:
                    pass

        # set chosen device for the live loop
        try:
            self._selected_device = chosen
        except Exception:
            pass

        # set sounddevice default to the chosen device so rec uses it reliably
        try:
            import sounddevice as _sd
            if chosen is not None:
                try:
                    _sd.default.device = chosen
                except Exception:
                    pass
        except Exception:
            pass

        # run the live transcript loop (it will handle None device cases)
        try:
            self._live_transcribe_loop()
        except Exception:
            pass

        # persist chosen device to config for next runs
        try:
            if chosen is not None:
                cfg_path = os.path.join(os.path.dirname(__file__), 'config.py')
                # read existing config
                try:
                    with open(cfg_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                except Exception:
                    lines = []
                out = []
                found = False
                for ln in lines:
                    if ln.strip().startswith('DEFAULT_MIC_DEVICE'):
                        out.append(f'DEFAULT_MIC_DEVICE = {repr(chosen)}\n')
                        found = True
                    else:
                        out.append(ln)
                if not found:
                    out.append('\n# persisted by UI auto-select\n')
                    out.append(f'DEFAULT_MIC_DEVICE = {repr(chosen)}\n')
                try:
                    with open(cfg_path, 'w', encoding='utf-8') as f:
                        f.writelines(out)
                    self.root.after(0, lambda: self.log_message('system', f'Persisted DEFAULT_MIC_DEVICE={chosen} to config.py'))
                except Exception:
                    pass
        except Exception:
            pass

    def _stop_live_listen(self):
        try:
            self._live_stop_event.set()
            self._listening = False
            try:
                self.mic_button.config(bg="#333333")
            except Exception:
                pass
            # stop recorder (writer finishes)
            self.recorder.stop()
            self.log_message('system', 'Live listening stopped.')
            # ensure thread joined briefly
            if self._live_thread is not None:
                self._live_thread.join(timeout=1)
            # send the full accumulated transcript on stop (conversation mode)
            final = (self._live_last_transcript or "").strip()
            if final:
                self.root.after(0, lambda: self.prompt_entry.delete(0, tk.END))
                self.root.after(0, lambda txt=final: self.prompt_entry.insert(0, txt))
                # send automatically the full conversation transcript
                threading.Thread(target=self._send_prompt_thread, args=(final, self.selected_model), daemon=True).start()
        except Exception as exc:
            self.log_message('system', f'Error stopping live listen: {exc}')

    def _live_transcribe_loop(self):
        """Background loop: periodically transcribe the recorder WAV and stream incremental words to the UI."""
        import tempfile
        import sounddevice as sd
        import numpy as np
        import wave

        samplerate = getattr(self.recorder, 'samplerate', 16000)
        channels = getattr(self.recorder, 'channels', 1)
        device = getattr(self.recorder, 'device', None)
        # prefer explicit selection from UI if present
        try:
            sel = getattr(self, '_selected_device', None)
            if sel is None:
                sel = None
        except Exception:
            sel = None
        if sel is not None:
            device = sel
        # use longer chunks for more stable ASR results
        chunk_seconds = 2.4
        # minimum characters in a transcription chunk before emitting incremental UI updates
        min_trans_chars = 6
        last = ""
        # debug: report device info
        try:
            # log chosen device index and name/details
            try:
                dev_info = sd.query_devices(device) if device is not None else sd.query_devices()
                name = dev_info.get('name') if isinstance(dev_info, dict) else str(dev_info)
                max_in = dev_info.get('max_input_channels', 'unknown') if isinstance(dev_info, dict) else 'unknown'
            except Exception:
                name = str(device)
                max_in = 'unknown'
            self.root.after(0, lambda: self.log_message('system', f'Live recorder device={device} ({name}), max_input_channels={max_in}, samplerate={samplerate}, channels={channels}'))
        except Exception:
            pass

        while not self._live_stop_event.is_set():
            try:
                # record a short chunk directly from the input device
                frames = int(samplerate * chunk_seconds)
                try:
                    # ensure we have a valid input device; if not, pick the first input-capable device
                    use_device = device
                    try:
                        if use_device is None:
                            dd = sd.default.device
                            if isinstance(dd, (list, tuple)):
                                use_device = dd[0]
                            else:
                                use_device = dd
                        info = sd.query_devices(use_device)
                        if info.get('max_input_channels', 0) <= 0:
                            # find first input-capable device
                            for i, d in enumerate(sd.query_devices()):
                                if d.get('max_input_channels', 0) > 0:
                                    use_device = i
                                    info = d
                                    break
                    except Exception:
                        use_device = None

                    # adapt samplerate/channels to the chosen device
                    try:
                        if use_device is not None:
                            dinfo = sd.query_devices(use_device)
                            dsr = int(dinfo.get('default_samplerate', samplerate))
                            dch = int(dinfo.get('max_input_channels', channels) or channels)
                            use_sr = dsr
                            use_ch = min(channels, dch) if dch > 0 else channels
                        else:
                            use_sr = samplerate
                            use_ch = channels
                    except Exception:
                        use_sr = samplerate
                        use_ch = channels

                    if use_device is None:
                        rec = sd.rec(frames, samplerate=use_sr, channels=use_ch, dtype='int16')
                    else:
                        try:
                            rec = sd.rec(frames, samplerate=use_sr, channels=use_ch, dtype='int16', device=use_device)
                        except Exception:
                            # try common fallback rates
                            rec = None
                            for try_sr in (44100, 48000, 16000):
                                try:
                                    rec = sd.rec(frames, samplerate=try_sr, channels=use_ch, dtype='int16', device=use_device)
                                    use_sr = try_sr
                                    break
                                except Exception:
                                    rec = None
                                    continue
                            if rec is None:
                                raise
                    sd.wait()
                except Exception as e:
                    self.root.after(0, lambda: self.log_message('system', f'Audio capture error: {e}'))
                    time.sleep(0.8)
                    continue

                # write to temp wav
                tmp_path = Path(tempfile.gettempdir()) / f"live_chunk_{int(time.time()*1000)}.wav"
                try:
                    with wave.open(str(tmp_path), 'wb') as wf:
                        wf.setnchannels(channels)
                        wf.setsampwidth(2)
                        wf.setframerate(samplerate)
                        wf.writeframes(rec.tobytes())
                except Exception:
                    try:
                        tmp_path.unlink()
                    except Exception:
                        pass
                    time.sleep(0.6)
                    continue

                # call async transcribe on the short chunk
                text = ""
                try:
                    future = asyncio.run_coroutine_threadsafe(self.ai.transcribe(str(tmp_path)), self.loop)
                    text = (future.result(timeout=10) or "").strip()
                except Exception:
                    text = ""

                # if remote transcription is empty or too short, try local Vosk fallback
                if not text or len(text.strip()) < 6:
                    # local ASR fallback removed — rely on remote transcription only
                    try:
                        self.root.after(0, lambda: self.log_message('system', 'No local ASR available; skipping fallback.'))
                    except Exception:
                        pass

                # cleanup temp file
                try:
                    tmp_path.unlink()
                except Exception:
                    pass

                # if transcription empty, skip
                if not text:
                    time.sleep(0.2)
                    continue

                # normalize whitespace
                text_norm = ' '.join(text.split())

                # ignore very short transcriptions (likely noise or single words)
                if len(text_norm) < min_trans_chars:
                    try:
                        self.root.after(0, lambda: self.log_message('system', f'Skipped short transcription (len={len(text_norm)})'))
                    except Exception:
                        pass
                    time.sleep(0.2)
                    continue

                # debug: compute RMS of recorded chunk
                try:
                    import numpy as _np
                    arr = _np.frombuffer(rec.tobytes(), dtype=_np.int16)
                    if arr.size:
                        rms = float(_np.sqrt(_np.mean((arr.astype('float32') / 32768.0) ** 2)))
                    else:
                        rms = 0.0
                except Exception:
                    rms = 0.0
                try:
                    self.root.after(0, lambda rms=rms: self.log_message('system', f'Chunk RMS={rms:.4f}, transcribed_len={len(text_norm)}'))
                except Exception:
                    pass

                # compute new words relative to last transcript
                if last and text_norm.startswith(last):
                    diff = text_norm[len(last):].strip()
                elif last and last in text_norm:
                    diff = text_norm.split(last, 1)[1].strip()
                else:
                    # big change, consider entire text as new
                    diff = text_norm

                last = text_norm
                self._live_last_transcript = text_norm

                if not diff:
                    time.sleep(0.1)
                    continue

                # split into words and emit only substantive tokens
                words = [w for w in diff.split() if len(w.strip(".,!?()[]{}\"'`")) > 1]
                if not words:
                    time.sleep(0.1)
                    continue

                new_chunk = ' '.join(words)

                def emit_chunk(c=new_chunk):
                    try:
                        cur = self.prompt_entry.get()
                        if cur and not cur.endswith(' '):
                            cur = cur + ' '
                        self.prompt_entry.delete(0, tk.END)
                        self.prompt_entry.insert(0, (cur + c).strip())
                    except Exception:
                        pass
                    # log the incremental words once
                    self.log_message('user', c)

                self.root.after(0, emit_chunk)

            except Exception:
                pass
            finally:
                time.sleep(0.15)

    def log_message(self, role: str, message: str):
        tag = "user" if role == "user" else "assistant" if role == "assistant" else "system"
        prefix = "You: " if role == "user" else "Assistant: " if role == "assistant" else "Info: "

        def append():
            start_index = self.log.index(tk.END)
            self.log.insert(tk.END, f"{prefix}{message}\n", tag)
            if role == "assistant":
                self.log.see(start_index)
            elif role == "user":
                # scroll to show new user message
                self.log.see(tk.END)
            # system messages do NOT force-scroll — user stays at the answer

        self.root.after(0, append)

    def run(self):
        self.root.mainloop()

    def _show_ocr_preview(self):
        """Show a small dialog with OCR text and options: Send, Retry (different PSM), Cancel."""
        data = getattr(self, '_last_ocr', None)
        if not data:
            return

        preview = tk.Toplevel(self.root)
        preview.title('OCR Preview')
        preview.geometry('640x360')
        preview.transient(self.root)

        text_box = tk.Text(preview, wrap='word')
        text_box.pack(fill='both', expand=True, padx=8, pady=8)
        text_box.insert('1.0', data['text'])

        btn_frame = tk.Frame(preview)
        btn_frame.pack(fill='x', pady=(0,8))

        def do_send():
            lang = getattr(self, 'lang_var', None).get() if getattr(self, 'lang_var', None) else 'Python'
            user_prompt = (
                "Act as a coding-interview assistant. Analyze the captured screen text and determine whether it contains a coding interview question or prompt. "
                "Respond using one of the two exact formats below (no extra commentary):\n\n"
                "If a coding question is present, reply exactly as:\n"
                "FOUND\n\n"
                "<Problem statement (concise)>\n\n"
                f"```{lang.lower()}\n<code solution>\n```\n\n"
                "If no coding question is present, reply exactly as:\n"
                "NOT_FOUND\n\n"
                "<one-line explanation why not>\n\n"
                f"Always prefer {lang} for solutions. Make code runnable and minimal. Now analyze the following captured text:\n\n"
                "Captured text:\n" + text_box.get('1.0', tk.END)
            )
            user_prompt += "\n\nAdditionally: if you include code, keep it concise, runnable, and focused on the requested solution. Do not add extra examples or long explanations unless asked."
            preview.destroy()
            self.log_message('system', 'Captured region; sending to AI...')
            threading.Thread(target=self._capture_ask_thread, args=(user_prompt, self.selected_model), daemon=True).start()

        def do_retry():
            # cycle to next PSM and rerun OCR
            idx = (data.get('psm_index', 0) + 1) % len(data['psm_list'])
            data['psm_index'] = idx
            psm = data['psm_list'][idx]
            try:
                import pytesseract
                custom = fr'--oem 1 --psm {psm}'
                new_text = pytesseract.image_to_string(data['img'], config=custom)
                data['text'] = new_text
                text_box.delete('1.0', tk.END)
                text_box.insert('1.0', new_text)
                self.log_message('system', f'Retried OCR with psm={psm}')
            except Exception as exc:
                self.log_message('system', f'OCR retry failed: {exc}')

        def do_cancel():
            preview.destroy()

        send_btn = tk.Button(btn_frame, text='Send', command=do_send, bg='#44b94e', fg='white')
        send_btn.pack(side='left', padx=6)
        retry_btn = tk.Button(btn_frame, text='Retry OCR (try other PSM)', command=do_retry, bg='#ffaa00')
        retry_btn.pack(side='left', padx=6)
        cancel_btn = tk.Button(btn_frame, text='Cancel', command=do_cancel, bg='#666666', fg='white')
        cancel_btn.pack(side='right', padx=6)

    def _preprocess_for_ocr(self, img):
        """Create several preprocessing variants for OCR to improve robustness.

        Returns a list of PIL images to try.
        Variants: grayscale+autocontrast, sharpened, binarized, inverted, small rotations to correct skew.
        """
        try:
            from PIL import ImageOps, ImageFilter, Image
        except Exception:
            return [img]

        ims = []
        base = img.convert('L')
        base = ImageOps.autocontrast(base)
        base = base.filter(ImageFilter.MedianFilter(size=3))
        try:
            base = base.resize((int(base.width * 1.5), int(base.height * 1.5)), Image.LANCZOS)
        except Exception:
            pass
        ims.append(base)

        # sharpened
        try:
            sh = base.filter(ImageFilter.SHARPEN)
            ims.append(sh)
        except Exception:
            pass

        # binarized (simple threshold) and inverted
        try:
            bw = base.point(lambda p: 255 if p > 140 else 0).convert('L')
            ims.append(bw)
            inv = ImageOps.invert(bw)
            ims.append(inv)
        except Exception:
            pass

        # try small rotations to correct skew
        angles = (-2, 0, 2)
        for a in angles:
            if a == 0:
                continue
            try:
                rot = base.rotate(a, expand=True)
                ims.append(rot)
            except Exception:
                pass

        # ensure uniqueness order preserved
        seen = []
        out = []
        for im in ims:
            key = (im.width, im.height)
            if key in seen:
                out.append(im)
            else:
                out.append(im)
                seen.append(key)
        return out

    def _ocr_best_psm(self, imgs, pytesseract, psm_list=(6,3,11,1)):
        """Try multiple preprocessed images and multiple PSMs; return the best-scoring text and psm.

        imgs: list of PIL images
        Returns (text, chosen_psm)
        """
        best = ('', None, -1)
        for im in imgs:
            for psm in psm_list:
                try:
                    cfg = fr'--oem 1 --psm {psm}'
                    txt = pytesseract.image_to_string(im, config=cfg)
                except Exception:
                    continue
                score = self._code_likelihood_score(txt)
                # small bonus for longer results to avoid empty short strings
                score += min(len(txt) // 50, 5)
                if score > best[2]:
                    best = (txt, psm, score)
        if best[1] is None:
            return ('', psm_list[0])
        return (best[0], best[1])

    def _normalize_capture_line(self, raw_line: str) -> str:
        line = ' '.join((raw_line or '').split())
        return line.strip(" |`~!@#$%^&*()_+-=[]{};:'\",.<>?/\\")

    def _extract_capture_task_text(self, text: str) -> str:
        """Extract the most task-like lines from noisy full-screen OCR."""
        if not text:
            return ""

        chrome_tokens = (
            'visual studio code', 'powershell', 'terminal', 'outline', 'timeline', 'problems',
            'this pc', 'recycle bin', 'type here to search', 'assistant:', 'info:', 'todos',
            'common_prompt.txt', 'gpt_4o_mini.txt', 'gpt_5.txt', 'ui.py', 'ai_bridge.py',
            'requirements.txt', 'readme', 'windows', 'utf-8', 'crlf', 'notepad', 'postman',
            'file edit format view help', 'ln 29, col 1', 'ln', 'col 1', 'untitled - notepad',
        )
        task_tokens = (
            'implement', 'design', 'build', 'create', 'write', 'solve', 'find', 'return',
            'given', 'array', 'string', 'linked list', 'tree', 'graph', 'sql', 'api',
            'latency', 'optimize', 'function', 'class', 'logger', 'module', 'system',
        )
        priority_tokens = ('implement', 'design', 'build', 'create', 'write', 'fix', 'add', 'make')

        normalized_lines = []
        for raw_line in text.splitlines():
            line = self._normalize_capture_line(raw_line)
            if line:
                normalized_lines.append(line)

        priority_lines = []
        for line in normalized_lines:
            lower = line.lower()
            if any(token in lower for token in chrome_tokens):
                continue
            if any(token in lower for token in priority_tokens) and sum(ch.isalpha() for ch in line) >= 10:
                priority_lines.append(line)

        if priority_lines:
            return '\n'.join(priority_lines[:3])[:1600]

        scored_lines = []
        for line in normalized_lines:
            if len(line) < 8:
                continue

            lower = line.lower()
            if any(token in lower for token in chrome_tokens):
                continue

            alpha_count = sum(ch.isalpha() for ch in line)
            if alpha_count < 6:
                continue

            score = 0
            if any(token in lower for token in task_tokens):
                score += 5
            if '?' in line or ':' in line:
                score += 2
            if len(line) >= 40:
                score += 2
            score += min(self._code_likelihood_score(line), 8)

            if score > 0:
                scored_lines.append((score, line))

        if not scored_lines:
            fallback_lines = []
            for line in normalized_lines:
                lower = line.lower()
                if any(token in lower for token in chrome_tokens):
                    continue
                if len(line) >= 20 and sum(ch.isalpha() for ch in line) >= 8:
                    fallback_lines.append(line)
                if len(fallback_lines) >= 6:
                    break
            return '\n'.join(fallback_lines)[:1600]

        top_lines = [line for _, line in sorted(scored_lines, key=lambda item: item[0], reverse=True)[:8]]
        seen = set()
        ordered = []
        for line in normalized_lines:
            if line in top_lines and line not in seen:
                ordered.append(line)
                seen.add(line)
        return '\n'.join(ordered)[:1600]

    def _capture_text_is_usable(self, candidate_text: str) -> bool:
        if not candidate_text:
            return False
        lower = candidate_text.lower()
        useful_tokens = (
            'implement', 'design', 'build', 'create', 'write', 'solve', 'find', 'return',
            'given', 'how', 'why', 'what', 'function', 'class', 'api', 'sql', 'logger',
            'latency', 'system', 'module', 'question', '?'
        )
        has_useful_token = any(token in lower for token in useful_tokens)
        if not has_useful_token:
            return False
        return len(candidate_text.strip()) >= 20

    def _capture_text_has_min_signal(self, text: str) -> bool:
        if not text:
            return False
        alpha_count = sum(ch.isalpha() for ch in text)
        digit_count = sum(ch.isdigit() for ch in text)
        return len(text.strip()) >= 60 and (alpha_count + digit_count) >= 25

    def _process_captured_image(self, img):
        """Background processing of a captured PIL image: OCR, logging, and scheduling send."""
        if self._has_active_request():
            self.root.after(0, lambda: self.log_message('system', 'A request is already running.'))
            return

        try:
            import pytesseract
        except Exception:
            self.root.after(0, lambda: self.log_message("system", "Capture failed: pytesseract not available. Install 'pytesseract' and ensure Tesseract OCR is installed on your system."))
            return

        try:
            if not self._configure_tesseract(pytesseract):
                self.root.after(0, lambda: self.log_message("system", "OCR failed: tesseract is not installed or not in PATH. See TESSERACT_INSTALLATION.md"))
                return
        except Exception:
            pass

        try:
            # Quick OCR fast-path: lightweight preprocessing and single PSM for minimal latency
            try:
                from PIL import ImageOps, Image
                quick = img.convert('L')
                quick = ImageOps.autocontrast(quick)
                try:
                    scale = 2.0 if max(quick.width, quick.height) >= 1400 else 1.5
                    quick = quick.resize((int(quick.width * scale), int(quick.height * scale)), Image.LANCZOS)
                except Exception:
                    pass
                cfg = r'--oem 1 --psm 6'
                quick_txt = pytesseract.image_to_string(quick, config=cfg)
                quick_conf = self._ocr_confidence(quick, pytesseract)
            except Exception:
                quick_txt = ''
                quick_conf = -1.0

            # Decide whether quick OCR is sufficient: require some alphanumeric density
            import re
            def _alnum_count(s: str) -> int:
                return len(re.findall(r"[A-Za-z0-9]", s or ""))

            if quick_txt and quick_txt.strip() and _alnum_count(quick_txt) >= 8 and quick_conf >= 45:
                # quick result appears reasonably dense; accept for low latency
                text = quick_txt
                chosen_psm = 6
                self.root.after(0, lambda: self.log_message('system', f'Quick OCR used psm={chosen_psm} (conf={quick_conf:.1f})'))
            else:
                proc_imgs = self._preprocess_for_ocr(img)
                text, chosen_psm = self._ocr_best_psm(proc_imgs, pytesseract, psm_list=[6, 3, 11, 1])
                self.root.after(0, lambda: self.log_message('system', f'OCR used psm={chosen_psm} after quick conf={quick_conf:.1f}'))
        except Exception as exc:
            self.root.after(0, lambda: self.log_message("system", f"OCR failed: {exc}"))
            return

        if not text or not text.strip():
            self.root.after(0, lambda: self.log_message('system', 'No text detected on screen.'))
            return

        # Log OCR text in main log and immediately send to AI for fast response
        self.root.after(0, lambda: self.log_message('system', 'Captured OCR text:'))
        self.root.after(0, lambda: self.log_message('system', text))

        candidate_text = self._extract_capture_task_text(text)
        if candidate_text:
            self.root.after(0, lambda: self.log_message('system', 'Focused task text:'))
            self.root.after(0, lambda: self.log_message('system', candidate_text))

        using_fallback_context = False
        if not self._capture_text_is_usable(candidate_text):
            if not self._capture_text_has_min_signal(text):
                self.root.after(0, lambda: self.log_message('assistant', 'SCREEN_CAPTURE_TOO_NOISY: Use region capture around the question or editor area only.'))
                return
            using_fallback_context = True
            candidate_text = text[:1200]
            self.root.after(0, lambda: self.log_message('system', 'Full-screen OCR is noisy; attempting a best-effort answer from raw capture text.'))

        # Check generic answers on focused text before hitting the API
        generic = self.ai.match_generic_answer(candidate_text)
        if generic:
            try:
                self.root.after(0, lambda: self.status_label.config(text="Answering now...", fg="#a6e22e"))
            except Exception:
                pass
            self.root.after(0, self._begin_assistant_stream)
            self.root.after(0, lambda g=generic: self._append_assistant_chunk(g))
            self.root.after(0, self._finish_assistant_stream)
            self.ai.db.save_message("user", candidate_text)
            self.ai.db.save_message("assistant", generic)
            return

        lang = getattr(self, 'lang_var', None).get() if getattr(self, 'lang_var', None) else 'Python'
        raw_excerpt = text[:1800]

        # Include any attached file in the capture prompt
        attached_section = ""
        attached_file = getattr(self, '_attached_file', None)
        if attached_file:
            file_content = self._read_file_content(attached_file)
            file_name = Path(attached_file).name
            if file_content and not file_content.startswith("[Error"):
                attached_section = f"\n\nAttached file ({file_name}):\n{file_content}\n--- End of attached file ---"
                self.root.after(0, lambda fn=file_name: self.log_message('system', f'Attached file included: {fn}'))
            self._clear_attached_file()

        # Build forced prompt for coding-interview interpretation (concise, runnable Python)
        forced = (
            "The OCR below comes from a noisy full-screen capture and may include editor tabs, file trees, terminal text, previous assistant output, timestamps, and unrelated desktop text. "
            "Ignore that noise and focus on the most likely user question or task. Use the focused task text first, then the raw OCR only as backup context. "
            f"If a clear task is present, answer it with a concise, runnable {lang} solution. "
            "If the task is vague, choose the simplest practical real-world interpretation first. Do not invent extra features such as thread safety, async behavior, rotation, multiple handlers, metadata fields, retries, or scalability unless the captured text asks for them. "
            "If there still is not enough signal, reply exactly with: SCREEN_CAPTURE_TOO_NOISY: Use region capture around the question.\n\n"
            f"Fallback mode: {'yes' if using_fallback_context else 'no'}\n\n"
            "Focused task text:\n" + candidate_text + "\n\n"
            "Raw OCR excerpt:\n" + raw_excerpt
            + attached_section
            + f"\n\nAdditionally: if you include code, keep it concise, runnable, and focused on the requested solution. Respond using {lang} code blocks where applicable. Do not add long explanations unless needed."
        )

        # start streaming response into UI with minimal latency
        try:
            self.root.after(0, lambda: self.status_label.config(text="Answering now...", fg="#a6e22e"))
        except Exception:
            pass
        try:
            self.root.after(0, self._begin_assistant_stream)
        except Exception:
            pass

        try:
            selected_model = self.selected_model
            future = asyncio.run_coroutine_threadsafe(
                self.ai.ask_gpt(forced, mode="chat", stream=True, on_delta=self._append_assistant_chunk, selected_model=selected_model, include_context=False),
                self.loop,
            )
            self._register_active_future(future)
            try:
                response = future.result(timeout=120)
                self.root.after(0, lambda response=response: self._hydrate_stream_buffer_from_result(response))
                self.root.after(0, self._finish_assistant_stream)
            except Exception as exc:
                self.root.after(0, self._finish_assistant_stream)
                if not self._is_cancelled_request_error(future, exc):
                    self.root.after(0, lambda exc=exc: self.log_message('assistant', f'Auto-send error: {type(exc).__name__}: {exc}'))
        except Exception as exc:
            self.root.after(0, self._finish_assistant_stream)
            self.root.after(0, lambda exc=exc: self.log_message('assistant', f'Auto-send failed: {exc}'))
        finally:
            self._clear_active_future(future if 'future' in locals() else None)
            try:
                self.root.after(0, lambda: self.status_label.config(text="Ready for the next prompt.", fg="#50fa7b"))
            except Exception:
                pass

    def _code_likelihood_score(self, text: str) -> int:
        """Heuristic scoring for how code-like OCR output is.

        Counts occurrences of code tokens and symbols.
        """
        if not text:
            return 0
        s = 0
        low = text.lower()
        keywords = ['def ', 'class ', 'import ', 'return ', 'if ', 'else:', 'elif ', 'for ', 'while ', 'lambda']
        symbols = ['(', ')', ':', '=', '->', '[', ']', '{', '}', '#']
        for kw in keywords:
            s += low.count(kw) * 5
        for sym in symbols:
            s += text.count(sym)
        # presence of '```' boosts score
        if '```' in text:
            s += 20
        # penalize very short outputs
        if len(text.strip()) < 20:
            s -= 5
        return s 




